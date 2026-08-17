"""Reports and export: one builder, three renderings."""
import datetime as dt
import io

import pytest
from openpyxl import load_workbook

from app.models.user import User, UserRole
from app.services import funding_reco, platform_analytics, report_formats, reports
from tests.conftest import auth_header, make_opportunity, make_profile, make_user

pypdf = pytest.importorskip("pypdf", reason="pypdf reads the generated PDFs back")

PATENT_STYLE_ID = "EP4672542A1"


@pytest.fixture
def seeded(client, db):
    """One of each role, plus a catalogue to report on."""
    owner = make_user(db, "owner@example.org", UserRole.RESEARCHER,
                      full_name="Professor Ada Lovelace")
    make_profile(db, owner, domains=["energy"], keywords=["battery", "storage"],
                 technology_areas=["energy storage"])
    make_user(db, "admin@example.org", UserRole.ADMIN, full_name="Platform Admin")
    make_user(db, "manager@example.org", UserRole.INNOVATION_MANAGER,
              full_name="Pipeline Manager")
    make_user(db, "quiet@example.org", UserRole.STARTUP_FOUNDER,
              full_name="Unset Founder")

    make_opportunity(db, title=f"Battery Storage Award {PATENT_STYLE_ID}",
                     domains=["energy"], keywords=["battery", "storage"],
                     amount_min=50_000, amount_max=502_434,
                     deadline=dt.date.today() + dt.timedelta(days=120))
    make_opportunity(db, title="Unrelated Astronomy Fund", domains=["astronomy"],
                     keywords=["telescope"])
    admin = auth_header(client, "admin@example.org")
    manager = auth_header(client, "manager@example.org")
    return {
        "funding": auth_header(client, "owner@example.org"),
        "system": admin,
        "accounts": admin,
        "pipeline": manager,
        "innovator": manager,
    }


DB_ONLY = ("funding", "system", "pipeline", "accounts")


def test_the_catalogue_offers_only_what_the_role_may_run(client, seeded):
    """Driven by the table the builder gates on."""
    def kinds(role_key):
        return {r["kind"] for r in client.get(
            "/api/reports", headers=seeded[role_key]).json()["reports"]}

    owner_kinds, admin_kinds, manager_kinds = (
        kinds("funding"), kinds("system"), kinds("pipeline"))

    assert "funding" in owner_kinds and "system" not in owner_kinds
    assert admin_kinds == {"system", "accounts"}
    assert manager_kinds == {"patents", "pipeline", "innovator"}
    for role, expected in ((UserRole.RESEARCHER, owner_kinds),
                           (UserRole.ADMIN, admin_kinds),
                           (UserRole.INNOVATION_MANAGER, manager_kinds)):
        assert expected == {k for k, v in reports.CATALOGUE.items()
                            if role in v["roles"]}


def test_every_offered_report_declares_whether_it_needs_a_subject(client, seeded):
    """A topic and an account are different questions."""
    for role_key in ("funding", "system", "pipeline"):
        for entry in client.get("/api/reports",
                                headers=seeded[role_key]).json()["reports"]:
            assert isinstance(entry["needs_subject"], bool), entry
            assert isinstance(entry["needs_query"], bool), entry
            assert not (entry["needs_subject"] and entry["needs_query"]), entry


def test_a_report_about_one_innovator_refuses_to_guess_which(client, seeded):
    missing = client.get("/api/reports/innovator", headers=seeded["innovator"])
    assert missing.status_code == 400
    assert "innovator" in missing.json()["detail"].lower()


def test_the_pipeline_report_does_not_re_score_the_roster(client, seeded, db):
    """One scoring path."""
    report = client.get("/api/reports/pipeline",
                        headers=seeded["pipeline"]).json()
    roster = {r["user_id"]: r for r in
              platform_analytics.pipeline_stats(db)["roster"]}
    printed = next(s for s in report["sections"] if s["heading"] == "Roster")

    match_column = printed["columns"].index("Best match")
    email_column = printed["columns"].index("Email")
    by_email = {u.email: u.id for u in db.query(User).all()}
    for row in printed["rows"]:
        entry = roster.get(by_email[row[email_column]])
        best = (entry or {}).get("best_match")
        expected = f"{round(best['score'])}%" if best else "—"
        assert row[match_column] == expected, row


def test_a_report_outside_the_role_is_refused_not_merely_hidden(client, seeded):
    assert client.get("/api/reports/system",
                      headers=seeded["funding"]).status_code == 403
    for caller in ("funding", "pipeline"):
        assert client.get("/api/reports/accounts",
                          headers=seeded[caller]).status_code == 403
    assert client.get("/api/reports/pipeline",
                      headers=seeded["system"]).status_code == 403


def test_the_accounts_report_prints_the_audit_log_the_screen_shows(client, seeded, db):
    """The audit table renders in one panel and can leave the platform nowhere else."""
    header = seeded["accounts"]
    target = db.query(User).filter(User.email == "quiet@example.org").first()
    assert client.put(f"/api/users/{target.id}/role", headers=header,
                      json={"role": "innovation_manager"}).status_code == 200

    on_screen = client.get("/api/users/audit", headers=header).json()
    printed = next(s for s in client.get("/api/reports/accounts", headers=header)
                   .json()["sections"] if s["heading"] == "Recent access changes")

    assert len(printed["rows"]) == len(on_screen)
    who = printed["columns"].index("Who")
    whom = printed["columns"].index("To whom")
    for row, event in zip(printed["rows"], on_screen):
        assert row[who] == event["actor_email"]
        assert row[whom] == event["target_email"]


def test_a_figure_that_needs_a_denominator_on_screen_carries_one_in_the_export(
        client, seeded):
    """Two figures were corrected on their cards and left bare in the PDF."""
    sections = {s["heading"]: s for s in
                client.get("/api/reports/system", headers=seeded["accounts"])
                      .json()["sections"]}

    reach = sections["Reach across portfolio owners"]
    if any(f["label"] == "Median best match" and f["value"] != "—"
           for f in reach["facts"]):
        assert reach["note"], "the median is printed with no population named"
        assert "matched" in reach["note"]

    cached = sections["Cached data"]
    figures = {f["label"]: f["value"] for f in cached["facts"]}
    both = int(figures["With a full corpus"]) + int(figures["Fallback only"])
    if both != int(figures["Topics cached"]):
        assert "do not sum" in cached["note"], (
            f"{figures['With a full corpus']} + {figures['Fallback only']} is not "
            f"{figures['Topics cached']}, and nothing says so")


def test_the_two_reports_of_one_roster_identify_people_the_same_way(client, seeded):
    """The manager's pipeline report and the admin's system report render the same."""
    printed = next(s for s in client.get("/api/reports/system", headers=seeded["accounts"])
                   .json()["sections"] if s["heading"] == "Pipeline")
    assert "Account" in printed["columns"] and "Email" in printed["columns"]
    account = printed["columns"].index("Account")
    for row in printed["rows"]:
        assert not str(row[account]).isdigit(), "a bare account id is not identification"


def test_an_unknown_report_is_a_404(client, seeded):
    assert client.get("/api/reports/nonsense",
                      headers=seeded["funding"]).status_code == 404


def _stated_figures(report: dict) -> set[str]:
    """Every figure a reader would see on screen, as the string they would see."""
    values = {str(item["value"]) for item in report["meta"]}
    for section in report["sections"]:
        values.update(str(fact["value"]) for fact in section.get("facts") or [])
    return {v for v in values if v and v != "—"}


def _sheet_values(book) -> tuple[set[str], set[float]]:
    text, numbers = set(), set()
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)):
                    numbers.add(float(cell.value))
                else:
                    text.add(str(cell.value))
    return text, numbers


@pytest.mark.parametrize("kind", DB_ONLY)
def test_the_spreadsheet_states_the_same_figures_as_the_screen(client, seeded, kind):
    headers = seeded[kind]
    on_screen = client.get(f"/api/reports/{kind}", headers=headers).json()
    export = client.get(f"/api/reports/{kind}?format=xlsx", headers=headers)
    assert export.status_code == 200

    text, numbers = _sheet_values(load_workbook(io.BytesIO(export.content)))
    for figure in _stated_figures(on_screen):
        value_kind, parsed = report_formats.classify(figure)
        if value_kind == "text":
            assert figure in text, f"{kind}: {figure!r} is on screen but not in the sheet"
        else:
            assert float(parsed) in numbers, (
                f"{kind}: the figure {figure!r} is on screen but "
                f"{parsed} is not in the sheet")


@pytest.mark.parametrize("kind", DB_ONLY)
def test_the_pdf_states_the_same_headings_and_subject(client, seeded, kind):
    headers = seeded[kind]
    on_screen = client.get(f"/api/reports/{kind}", headers=headers).json()
    export = client.get(f"/api/reports/{kind}?format=pdf", headers=headers)
    assert export.status_code == 200
    assert export.content.startswith(b"%PDF")

    text = _pdf_text(export.content)
    assert on_screen["title"] in text
    for section in on_screen["sections"]:
        assert section["heading"] in text, f"{kind}: section {section['heading']} missing"


def _pdf_text(body: bytes) -> str:
    reader = pypdf.PdfReader(io.BytesIO(body))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def test_an_identifier_is_never_broken_across_a_line(client, seeded):
    """Column widths take their minimum from the widest unbreakable token."""
    body = client.get("/api/reports/funding?format=pdf",
                      headers=seeded["funding"]).content
    text = _pdf_text(body)
    assert PATENT_STYLE_ID in text, (
        f"{PATENT_STYLE_ID} did not survive the layout whole:\n{text[:600]}")


def test_a_shortened_title_ends_on_a_word(client, db, seeded):
    """A hard slice produced "power tool and power tool sy", which reads as damage."""
    long_title = ("Advanced Grid Scale Battery Storage Demonstration Programme "
                  "For Distribution Networks")
    make_opportunity(db, title=long_title, domains=["energy"],
                     keywords=["battery", "storage"])
    on_screen = client.get("/api/reports/funding", headers=seeded["funding"]).json()
    best = next(f["value"] for f in on_screen["sections"][0]["facts"]
                if f["label"] == "Best match grant")
    if best.endswith("…"):
        assert not best[:-1].endswith(" ")
        assert long_title.startswith(best[:-1]) or best[:-1] in long_title
        assert best[:-1].split()[-1] in long_title.split()


def test_every_number_in_the_sheet_carries_a_display_format(client, seeded):
    """Stored as numbers so they sort and total; formatted so they stay readable."""
    export = client.get("/api/reports/system?format=xlsx", headers=seeded["system"])
    book = load_workbook(io.BytesIO(export.content))
    checked = 0
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    assert cell.number_format != "General", (
                        f"{sheet.title}!{cell.coordinate} is a bare number")
                    assert cell.alignment.horizontal == "right", (
                        f"{sheet.title}!{cell.coordinate} is not aligned as a figure")
                    checked += 1
    assert checked, "the sheet contained no numbers to check"


def test_thousands_are_grouped_and_percentages_marked(client, seeded):
    export = client.get("/api/reports/funding?format=xlsx", headers=seeded["funding"])
    book = load_workbook(io.BytesIO(export.content))
    formats = {c.number_format
               for sheet in book.worksheets for row in sheet.iter_rows() for c in row
               if isinstance(c.value, (int, float))}
    assert any("#,##0" in f for f in formats)
    assert any("%" in f for f in formats), "match scores are percentages"


def test_a_ruled_out_grant_is_not_counted_as_a_strong_match(client, db, seeded):
    """The report once claimed a strong match while naming a 17% grant as the best."""
    make_opportunity(db, title="Closed Battery Programme", domains=["energy"],
                     keywords=["battery", "storage"],
                     deadline=dt.date.today() - dt.timedelta(days=1))
    on_screen = client.get("/api/reports/funding", headers=seeded["funding"]).json()
    facts = {f["label"]: f["value"] for f in on_screen["sections"][0]["facts"]}

    strong_label = f"Strong matches (≥{platform_analytics.STRONG_MATCH}%)"
    strong = int(facts[strong_label].replace(",", ""))
    best = int(facts["Best match"].rstrip("%"))
    if strong:
        assert best >= platform_analytics.STRONG_MATCH, (
            "a strong match is claimed but the best match named is below the floor")

    headings = {s["heading"] for s in on_screen["sections"]}
    assert "Ruled out" in headings or "Closed Battery Programme" not in str(on_screen)


def test_dates_read_the_same_way_everywhere(client, seeded):
    """A funding deadline printed `05 Dec 2026` while a patent date printed."""
    assert reports._date(dt.date(2026, 12, 5)) == "05 Dec 2026"
    assert reports._date("2025-12-31") == "31 Dec 2025"
    assert reports._date("2025-12-31T09:00:00Z") == "31 Dec 2025"
    assert reports._date(None) == "—"


def test_a_download_names_itself(client, seeded):
    for extension in ("xlsx", "pdf"):
        response = client.get(f"/api/reports/funding?format={extension}",
                              headers=seeded["funding"])
        disposition = response.headers["content-disposition"]
        assert disposition.startswith("attachment;")
        assert f".{extension}" in disposition
        assert dt.date.today().isoformat() in disposition


def test_an_unsupported_format_is_rejected_by_the_route(client, seeded):
    assert client.get("/api/reports/funding?format=docx",
                      headers=seeded["funding"]).status_code == 422


def test_the_report_names_who_it_was_prepared_for(client, seeded):
    on_screen = client.get("/api/reports/funding", headers=seeded["funding"]).json()
    prepared = {m["label"]: m["value"] for m in on_screen["meta"]}
    assert prepared["Prepared for"] == "Professor Ada Lovelace"
    assert prepared["Role"] == "Researcher"
    assert PATENT_STYLE_ID in _pdf_text(
        client.get("/api/reports/funding?format=pdf",
                   headers=seeded["funding"]).content)


def test_a_report_needs_the_profile_it_reports_on(client, db):
    """An owner with no portfolio gets a stated reason, not an empty document."""
    make_user(db, "bare@example.org", UserRole.RESEARCHER)
    response = client.get("/api/reports/funding",
                          headers=auth_header(client, "bare@example.org"))
    assert response.status_code == 400
    assert "profile" in response.json()["detail"].lower()


def test_the_classifier_tells_figures_from_words(client):
    """The one decision both renderers depend on: right-align or not, format or not."""
    assert report_formats.classify("502,434")[0] == "int"
    assert report_formats.classify("27%")[0] == "pct"
    assert report_formats.classify("5.2")[0] == "float"
    assert report_formats.classify("—")[0] == "text"
    assert report_formats.classify(PATENT_STYLE_ID)[0] == "text"
    assert report_formats.classify("2026-12-05")[0] == "text"
