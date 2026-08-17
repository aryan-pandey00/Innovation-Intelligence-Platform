"""Rendering a report to a spreadsheet and to a PDF."""
from __future__ import annotations

import io
import re
import unicodedata

import os

import reportlab
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
    TableStyle,
)


def _register_fonts() -> tuple[str, str, str, set[int]]:
    """A font that can draw the names in the data."""
    folder = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
    try:
        regular = TTFont("Vera", os.path.join(folder, "Vera.ttf"))
        bold = TTFont("Vera-Bold", os.path.join(folder, "VeraBd.ttf"))
        italic = TTFont("Vera-Italic", os.path.join(folder, "VeraIt.ttf"))
        for face in (regular, bold, italic):
            pdfmetrics.registerFont(face)
        pdfmetrics.registerFontFamily("Vera", normal="Vera", bold="Vera-Bold",
                                      italic="Vera-Italic")
        return "Vera", "Vera-Bold", "Vera-Italic", set(regular.face.charToGlyph)
    except Exception:
        return "Helvetica", "Helvetica-Bold", "Helvetica-Oblique", set()


_FONT, _FONT_BOLD, _FONT_ITALIC, _GLYPHS = _register_fonts()

_NAVY = colors.HexColor("#12203c")
_MUTED = colors.HexColor("#64748b")
_LINE = colors.HexColor("#e2e8f0")
_SURFACE = colors.HexColor("#f8fafc")

_ACCENT = colors.HexColor("#cca01a")

_CELL_PT = 8.4
_HEAD_PT = 7.6
_CELL_PAD = 15

_XL_HEAD = PatternFill("solid", fgColor="12203C")
_XL_HEAD_FONT = Font(bold=True, color="FFFFFF")

_SHEET_BAD = re.compile(r"[:\\/?*\[\]]")


def _sheet_name(heading: str, used: set[str]) -> str:
    name = _SHEET_BAD.sub(" ", heading).strip()[:31] or "Section"
    if name in used:
        base = name[:28]
        n = 2
        while f"{base} {n}" in used:
            n += 1
        name = f"{base} {n}"
    used.add(name)
    return name


def classify(value) -> tuple[str, object]:
    """What a formatted value actually is: ("int"|"float"|"pct"|"text", parsed)."""
    if value is None:
        return "text", ""
    if isinstance(value, bool):
        return "text", str(value)
    if isinstance(value, int):
        return "int", value
    if isinstance(value, float):
        return ("int", int(value)) if value.is_integer() else ("float", value)

    text = str(value).strip()
    if text in ("—", "", "-"):
        return "text", ""
    bare = text.replace(",", "")
    percent = bare.endswith("%")
    if percent:
        bare = bare[:-1]
    bare = bare.lstrip("+")
    try:
        number = float(bare)
    except ValueError:
        return "text", text
    if percent:
        return "pct", number
    return ("int", int(number)) if number.is_integer() else ("float", number)


_XL_FORMAT = {"int": "#,##0", "float": "#,##0.0", "pct": '#,##0.0"%"'}


def _as_cell(value):
    """The value a spreadsheet cell should hold, without its display format."""
    return classify(value)[1]


def _write(ws, row: int, column: int, value) -> None:
    """One cell, with the number format its type deserves."""
    kind, parsed = classify(value)
    cell = ws.cell(row=row, column=column, value=parsed)
    if kind in _XL_FORMAT:
        cell.number_format = _XL_FORMAT[kind]
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.alignment = Alignment(vertical="top", wrap_text=len(str(parsed)) > 60)


def to_xlsx(report: dict) -> bytes:
    wb = Workbook()
    used: set[str] = set()

    cover = wb.active
    cover.title = _sheet_name("Report", used)
    cover["A1"] = report["title"]
    cover["A1"].font = Font(bold=True, size=16, color="12203C")
    cover["A2"] = report.get("summary", "")
    cover["A2"].font = Font(italic=True, color="64748B")
    row = 4
    for item in report.get("meta", []):
        cover.cell(row=row, column=1, value=item["label"]).font = Font(bold=True)
        cover.cell(row=row, column=2, value=str(item["value"]))
        row += 1
    cover.column_dimensions["A"].width = 26
    cover.column_dimensions["B"].width = 62
    cover["A2"].alignment = Alignment(wrap_text=True)

    for section in report.get("sections", []):
        ws = wb.create_sheet(_sheet_name(section["heading"], used))
        r = 1
        ws.cell(row=r, column=1, value=section["heading"]).font = Font(
            bold=True, size=13, color="12203C")
        r += 1
        if section.get("note"):
            cell = ws.cell(row=r, column=1, value=section["note"])
            cell.font = Font(italic=True, color="64748B", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            span = max(len(section.get("columns") or []), 4)
            ws.merge_cells(start_row=r, start_column=1, end_row=r,
                           end_column=span)
            lines = max(1, -(-len(section["note"]) // (span * 15)))
            ws.row_dimensions[r].height = 14 * lines + 4
            r += 2

        if section.get("facts"):
            ws.cell(row=r, column=1, value="Figure").font = _XL_HEAD_FONT
            ws.cell(row=r, column=1).fill = _XL_HEAD
            ws.cell(row=r, column=2, value="Value").font = _XL_HEAD_FONT
            ws.cell(row=r, column=2).fill = _XL_HEAD
            r += 1
            for fact in section["facts"]:
                ws.cell(row=r, column=1, value=fact["label"])
                _write(ws, r, 2, fact["value"])
                r += 1
            r += 1

        columns = section.get("columns") or []
        if columns:
            for c, name in enumerate(columns, start=1):
                cell = ws.cell(row=r, column=c, value=name)
                cell.font = _XL_HEAD_FONT
                cell.fill = _XL_HEAD
            head_row = r
            r += 1
            for data_row in section.get("rows") or []:
                for c, value in enumerate(data_row, start=1):
                    _write(ws, r, c, value)
                r += 1
            if section.get("rows"):
                ws.auto_filter.ref = (
                    f"A{head_row}:{get_column_letter(len(columns))}{r - 1}")
            ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

        _fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fit_columns(ws, limit: int = 60) -> None:
    """Width from the longest value, capped so one long abstract cannot dominate."""
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or isinstance(cell, type(None)):
                continue
            if not hasattr(cell, "column") or cell.column is None:
                continue
            length = len(str(cell.value))
            if length and length < 200:
                widths[cell.column] = max(widths.get(cell.column, 10), length + 2)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(width, limit)


def _pdf_styles() -> dict:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "RTitle", parent=base["Title"], fontName=_FONT_BOLD, fontSize=20,
            leading=24, textColor=_NAVY, alignment=TA_LEFT, spaceAfter=2),
        "subtitle": ParagraphStyle(
            "RSub", parent=base["Normal"], fontName=_FONT_ITALIC, fontSize=10.5,
            leading=15, textColor=_MUTED, spaceAfter=14),
        "heading": ParagraphStyle(
            "RHead", parent=base["Heading2"], fontName=_FONT_BOLD, fontSize=13,
            leading=17, textColor=_NAVY, spaceBefore=16, spaceAfter=4),
        "note": ParagraphStyle(
            "RNote", parent=base["Normal"], fontName=_FONT_ITALIC, fontSize=8.6,
            leading=12.4, textColor=_MUTED, spaceAfter=8),
        "metalabel": ParagraphStyle(
            "RMetaLabel", parent=base["Normal"], fontName=_FONT, fontSize=6.8,
            leading=9, textColor=_MUTED, spaceAfter=1),
        "metavalue": ParagraphStyle(
            "RMetaValue", parent=base["Normal"], fontName=_FONT, fontSize=9.4,
            leading=12.5, textColor=_NAVY),
        "cell": ParagraphStyle(
            "RCell", parent=base["Normal"], fontName=_FONT, fontSize=8.4,
            leading=11.6, textColor=_NAVY),
        "cellnum": ParagraphStyle(
            "RCellNum", parent=base["Normal"], fontName=_FONT, fontSize=8.4,
            leading=11.6, textColor=_NAVY, alignment=TA_RIGHT),
        "cellhead": ParagraphStyle(
            "RCellHead", parent=base["Normal"], fontSize=7.6, leading=10.4,
            textColor=colors.white, fontName=_FONT_BOLD),
        "cellheadnum": ParagraphStyle(
            "RCellHeadNum", parent=base["Normal"], fontSize=7.6, leading=10.4,
            textColor=colors.white, fontName=_FONT_BOLD, alignment=TA_RIGHT),
        "factlabel": ParagraphStyle(
            "RFactLabel", parent=base["Normal"], fontName=_FONT, fontSize=7.2,
            leading=9.5, textColor=_MUTED, spaceAfter=1),
        "factvalue": ParagraphStyle(
            "RFactValue", parent=base["Normal"], fontName=_FONT_BOLD, fontSize=12.5,
            leading=15.5, textColor=_NAVY),
        "factvalue_sm": ParagraphStyle(
            "RFactValueSm", parent=base["Normal"], fontName=_FONT_BOLD, fontSize=9.6,
            leading=13, textColor=_NAVY),
    }


def _stat_grid(items: list[dict], styles: dict, width: float, *,
               columns: int, label_style: str, value_style: str,
               gap: float = 6) -> Table:
    """Figures laid out as a grid of caption-above-value tiles."""
    if not items:
        return Table([[""]], colWidths=[width])
    cells = []
    small = f"{value_style}_sm"
    for item in items:
        text = _escape(item["value"])
        style = small if len(str(item["value"])) > 20 and small in styles else value_style
        cells.append([Paragraph(_escape(item["label"]).upper(), styles[label_style]),
                      Paragraph(text, styles[style])])
    rows = []
    for i in range(0, len(cells), columns):
        chunk = cells[i:i + columns]
        while len(chunk) < columns:
            chunk.append([Paragraph("", styles[label_style]),
                          Paragraph("", styles[value_style])])
        rows.append(chunk)

    flat = [[Table([[c[0]], [c[1]]], colWidths=[width / columns - gap],
                   style=TableStyle([
                       ("LEFTPADDING", (0, 0), (-1, -1), 0),
                       ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                       ("TOPPADDING", (0, 0), (-1, -1), 0),
                       ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
                       ("BOTTOMPADDING", (0, 1), (-1, 1), 0),
                   ])) for c in row] for row in rows]

    table = Table(flat, colWidths=[width / columns] * columns, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), gap),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, _LINE),
    ]))
    return table


def _rule(width: float, colour, thickness: float) -> Table:
    """A drawn line."""
    rule = Table([[""]], colWidths=[width], rowHeights=[thickness])
    rule.hAlign = "LEFT"
    rule.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), thickness, colour),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return rule


def _meta_grid(items: list[dict], styles: dict, width: float) -> Table:
    return _stat_grid(items, styles, width, columns=4,
                      label_style="metalabel", value_style="metavalue")


def _fact_grid(facts: list[dict], styles: dict, width: float) -> Table:
    return _stat_grid(facts, styles, width, columns=3,
                      label_style="factlabel", value_style="factvalue")


def _column_kinds(columns: list, rows: list[list]) -> list[bool]:
    """Which columns hold figures, so they can be set right-aligned."""
    numeric = []
    for i in range(len(columns)):
        values = [row[i] for row in rows if i < len(row)]
        filled = [v for v in values if str(v).strip() not in ("", "—", "-")]
        if not filled:
            numeric.append(False)
            continue
        hits = sum(1 for v in filled if classify(v)[0] in ("int", "float", "pct"))
        numeric.append(hits >= max(1, len(filled) * 0.6))
    return numeric


def _column_widths(columns: list, rows: list[list], numeric: list[bool],
                   width: float) -> list[float]:
    """Share the page out by how much text each column actually carries."""
    demand = []
    for i, name in enumerate(columns):
        values = [str(row[i]) for row in rows if i < len(row)]
        longest = max([len(v) for v in values] + [len(str(name))])
        typical = (sum(len(v) for v in values) / len(values)) if values else 0
        want = max(len(str(name)) + 2, typical * 1.15, min(longest, 46) * 0.6)
        demand.append(min(want, 14) if numeric[i] else want)
    total = sum(demand) or 1
    natural = total * 5.6
    usable = min(width, max(width * 0.45, natural))

    mins = []
    for i, name in enumerate(columns):
        body = [t for row in rows if i < len(row) for t in (str(row[i]).split() or [""])]
        head = str(name).split() or [""]
        widest = max(
            [pdfmetrics.stringWidth(t[:24], _FONT, _CELL_PT) for t in body]
            + [pdfmetrics.stringWidth(t[:24], _FONT_BOLD, _HEAD_PT) for t in head]
            or [0])
        mins.append(widest + _CELL_PAD)
    if sum(mins) > usable:
        squeeze = usable / sum(mins)
        mins = [m * squeeze for m in mins]

    widths = [usable * d / total for d in demand]
    for _ in range(4):
        deficit = sum(max(0.0, mins[i] - widths[i]) for i in range(len(widths)))
        if deficit < 0.5:
            break
        slack = sum(max(0.0, widths[i] - mins[i]) for i in range(len(widths)))
        if slack <= 0:
            break
        take = min(deficit, slack)
        widths = [w - (max(0.0, w - mins[i]) / slack) * take
                  for i, w in enumerate(widths)]
        widths = [max(widths[i], mins[i]) for i in range(len(widths))]

    scale = usable / sum(widths)
    return [w * scale for w in widths]


def _data_table(section: dict, styles: dict, width: float) -> Table:
    columns = section["columns"]
    rows = section["rows"]
    numeric = _column_kinds(columns, rows)

    head = [Paragraph(_escape(c),
                      styles["cellheadnum"] if numeric[i] else styles["cellhead"])
            for i, c in enumerate(columns)]
    body = [[Paragraph(_escape(v),
                       styles["cellnum"] if numeric[i] else styles["cell"])
             for i, v in enumerate(row)] for row in rows]

    table = Table([head] + body, colWidths=_column_widths(columns, rows, numeric, width),
                  repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, _LINE),
        ("LINEBELOW", (0, 0), (-1, 0), 0, colors.transparent),
    ]
    for i, is_numeric in enumerate(numeric):
        if is_numeric:
            style.append(("ALIGN", (i, 0), (i, -1), "RIGHT"))
    for i in range(2, len(body) + 1, 2):
        style.append(("BACKGROUND", (0, i), (-1, i), _SURFACE))
    table.setStyle(TableStyle(style))
    return table


def _escape(value) -> str:
    """Make a value safe to draw: valid mini-HTML, and no glyph the font lacks."""
    text = "" if value is None else str(value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    if not _GLYPHS or all(ord(ch) in _GLYPHS for ch in text):
        return text
    out = []
    for ch in text:
        if ord(ch) in _GLYPHS:
            out.append(ch)
            continue
        folded = unicodedata.normalize("NFKD", ch)
        kept = "".join(c for c in folded if ord(c) in _GLYPHS)
        out.append(kept or "?")
    rendered = "".join(out)
    if rendered and set(rendered.replace(" ", "")) == {"?"}:
        return "(not in Latin script)"
    return rendered


def to_pdf(report: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"{report['title']} — {report.get('subject', '')}".strip(" —"),
        author="Innovation Intelligence Platform",
    )
    width = doc.width
    styles = _pdf_styles()
    story = [
        Paragraph(_escape(report["title"]), styles["title"]),
        Paragraph(_escape(report.get("summary", "")), styles["subtitle"]),
        _rule(width * 0.16, _ACCENT, 2.2),
        Spacer(1, 12),
        _meta_grid(report.get("meta", []), styles, width),
        Spacer(1, 2),
    ]

    for section in report.get("sections", []):
        block = [Paragraph(_escape(section["heading"]), styles["heading"])]
        if section.get("note"):
            block.append(Paragraph(_escape(section["note"]), styles["note"]))
        if section.get("facts"):
            block.append(_fact_grid(section["facts"], styles, width))
            block.append(Spacer(1, 6))
        if section.get("columns"):
            if section.get("rows"):
                story.append(KeepTogether(block))
                story.append(_data_table(section, styles, width))
                continue
            block.append(Paragraph("Nothing to report in this section.",
                                   styles["note"]))
        story.append(KeepTogether(block))

    doc.build(story, onFirstPage=_page_furniture, onLaterPages=_page_furniture)
    return buffer.getvalue()


def _page_furniture(canvas, doc) -> None:
    """A rule and a page number, so a printed page is identifiable on its own."""
    canvas.saveState()
    canvas.setStrokeColor(_LINE)
    canvas.setLineWidth(0.5)
    y = 12 * mm
    canvas.line(18 * mm, y + 5 * mm, A4[0] - 18 * mm, y + 5 * mm)
    canvas.setFont(_FONT, 8)
    canvas.setFillColor(_MUTED)
    canvas.drawString(18 * mm, y, "Innovation Intelligence Platform")
    canvas.drawRightString(A4[0] - 18 * mm, y, f"Page {doc.page}")
    canvas.restoreState()
